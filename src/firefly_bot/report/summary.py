"""Per-run audit report: one row per processed document, with a hyperlink to each transaction.

Written as .xlsx (openpyxl). The transaction column is a clickable hyperlink into the Firefly
UI so you can review/edit auto-written entries in one click.
"""

from __future__ import annotations

import os
from datetime import datetime, timezone
from typing import Protocol

from openpyxl import Workbook
from openpyxl.styles import Font

from firefly_bot.models import MatchResult


class ReportWriter(Protocol):
    def write(self, results: list[MatchResult], report_dir: str) -> str: ...


class XlsxReportWriter:
    """Default report writer — emits the .xlsx audit file."""

    def write(self, results: list[MatchResult], report_dir: str) -> str:
        return write_xlsx(results, report_dir)

_HEADERS = (
    "Document",
    "Total",
    "Currency",
    "Counterparty IBAN",
    "Outcome",
    "Score",
    "Transaction",
    "Detail",
)


def write_xlsx(results: list[MatchResult], report_dir: str) -> str:
    """Write the report and return the file path."""
    os.makedirs(report_dir, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    path = os.path.join(report_dir, f"firefly-bot-{stamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Run summary"
    ws.append(list(_HEADERS))
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for result in results:
        inv = result.invoice
        ws.append(
            [
                inv.source.filename,
                float(inv.total_amount) if inv.total_amount is not None else None,
                inv.currency,
                inv.counterparty_iban or "",
                result.outcome.value,
                result.score,
                result.transaction_web_url or "",
                result.detail,
            ]
        )
        url = result.transaction_web_url
        if url:
            link_cell = ws.cell(row=ws.max_row, column=7)
            link_cell.hyperlink = url
            link_cell.value = result.transaction.id if result.transaction else url
            link_cell.font = Font(color="0000EE", underline="single")

    wb.save(path)
    return path
